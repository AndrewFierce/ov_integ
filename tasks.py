import io
import json
import uuid
from datetime import datetime
from uuid import uuid4

import pandas as pd
from celery.utils.log import get_task_logger
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.contrib.postgres.aggregates import StringAgg
from django.db.models import Subquery, OuterRef, Value, CharField, Q, F
from django.db.models.functions import Concat
from openpyxl.reader.excel import load_workbook

from big3_data_main_app.celery import app
from integrations.ched_integration.utils import ChedRequest
from file_builder.tools.tools_print_form import get_model_class
from file_upload.models import UploadedFile
from integrations.ovga_integration.models import PassRequests, TypeDocument, DriverDocument, TransportDocument
from model_app.vehicles.models import Transport
from notification.models import Notice
from notification import constants as c
from django.contrib.auth.models import User
from model_app.base.models import User, ContactType, Contact
from workflow.models import DocType

req = ChedRequest()
module_name = 'integrations.ovga_integration.tasks'


def send_notice(document_name, error, user_id):
    Notice.send_notice(
        User.objects.get(id=user_id),
        operation_type=c.OPERATION_INFO,
        content=f"Загрузка {document_name} произошла с ошибкой: {error}",
        theme=c.THEME_ERROR
    )


def check_load_sts_pts(wb, file_upload, transport, document_transport, user_id, doc_status_id):
    if 'СТС' in wb.sheetnames and 'ПТС' in wb.sheetnames:
        from integrations.ovga_integration.utils import load_sts_pts
        try:
            load_sts_pts(file_upload, transport, document_transport, doc_status_id)
        except Exception as e:
            send_notice('СТС/ПТС', e.args.__str__(), user_id)


def check_load_diagnostic_card(wb, file_upload, transport, document_transport, user_id, doc_status_id):
    from integrations.ovga_integration.utils import load_diag_card, diagnostic_card_str
    if diagnostic_card_str in wb.sheetnames:
        try:
            load_diag_card(file_upload, transport, document_transport, doc_status_id)
        except Exception as e:
            send_notice('Диагностических карт', e.args.__str__(), user_id)


def check_load_lizing(wb, file_upload, transport, document_transport, user_id, doc_status_id):
    from integrations.ovga_integration.utils import lizing_provider_str, load_liz
    if lizing_provider_str in wb.sheetnames:
        try:
            load_liz(file_upload, transport, document_transport, doc_status_id)
        except Exception as e:
            send_notice('Лизингодателей', e.args.__str__(), user_id)


def check_load_drivers(wb, file_upload, user_id, doc_status_id):
    from integrations.ovga_integration.utils import drivers_str, load_drivers
    if drivers_str in wb.sheetnames:
        try:
            load_drivers(file_upload, doc_status_id)
        except Exception as e:
            send_notice('Водителей', e.args.__str__(), user_id)


@app.task
def load_data_transport_documents(file_id=None, user_id=None):
    from workflow.models import DocStatus
    import pandas as pd
    from integrations.ovga_integration.models import TransportDocument
    from model_app.vehicles.models import Transport
    logger = get_task_logger("commands")
    logger.info('Получаем все TransportModel, TransportBrand, Transport')
    transport = pd.DataFrame(Transport.objects.exclude(deleted=True).values('id', 'tracker_id'))
    transport.rename(columns={'id': 'transport_id'}, inplace=True)
    document_transport = pd.DataFrame(TransportDocument.objects
                                      .values('id', 'serial_number', 'number', 'date_end', 'types_of_document_id',
                                              'issue_organization', 'transport__tracker_id'))
    if document_transport.empty:
        document_transport = pd.DataFrame(
            columns=['id', 'serial_number', 'number', 'date_end', 'issue_organization', 'types_of_document_id',
                     'transport__tracker_id'])
    document_transport.rename(columns={'transport__tracker_id': 'tracker_id'}, inplace=True)

    file_upload = UploadedFile.objects.get(id=file_id).file

    wb = load_workbook(file_upload, read_only=True)

    doc_status_id = DocStatus.objects.get_or_create(code='novyj-status', defaults={'name': 'Новая'})[0].id

    check_load_sts_pts(wb, file_upload, transport, document_transport, user_id, doc_status_id)

    check_load_diagnostic_card(wb, file_upload, transport, document_transport, user_id, doc_status_id)

    check_load_lizing(wb, file_upload, transport, document_transport, user_id, doc_status_id)

    check_load_drivers(wb, file_upload, user_id, doc_status_id)

    logger.info(f'Завершено, документы загружены в БД')
    if user_id:
        Notice.send_notice(
            User.objects.get(id=user_id),
            operation_type=c.OPERATION_INFO,
            content="Загрузка документов транспорта и водителей завершена...",
            theme=c.THEME_SHOW
        )
    return True


def sign_append(sign_list, doc_id):
    status_list = []
    for sign in sign_list:
        if sign is None or sign.__len__() < 50:
            continue
        sign = sign.replace('\r', '').replace('\n', '').replace(' ', '')
        sign_file = req.document_sign(doc_id, sign, 'Подпись файла')
        if sign_file.status_code != 200 or not sign_file.text:
            if sign is not None:
                sign = sign[:10] + '...' + sign[-10:]
            else:
                sign = 'подпись отсутствует'
            status_list.append(sign)
    return status_list


def notice_send(obj, user_id, is_success, res_file_text, theme=c.THEME_WARNING):
    if obj:
        Notice.send_ws_by_object(obj)
    Notice.send_notice(
        User.objects.get(id=user_id),
        operation_type=c.OPERATION_INFO,
        theme=theme,
        is_success=is_success,
        content=res_file_text,
    )


@app.task
def send_file_sign_ched(model_name, obj_id, document_id, status_id, status_to_id, document_class, asguf_code, mimetype, sign_list, user_id):
    is_success = False
    model_class, _ = get_model_class(model_name)
    obj = model_class.objects.get(id=obj_id)
    obj.status_id = status_id
    if not document_id:
        notice_send(obj, user_id, is_success, "Необходимо загрузить скан документа.")
        obj.save()
        return False
    document = UploadedFile.objects.get(id=document_id)
    if obj.doc_id:
        req.document_delete(obj.doc_id.__str__())
    obj.task_id = None
    res_file = req.document_send(document, document_class, document.name, asguf_code, mimetype)
    if res_file.status_code == 200:
        obj.doc_id = res_file.text
        if not sign_list or all(item is None for item in sign_list):
            notice_message = 'Не найдена подпись документа. Подпишите скан документа.'
            notice_send(obj, user_id, is_success, notice_message)
            obj.save()
            return False
        status_list = sign_append(sign_list, obj.doc_id.__str__())
        if status_list:
            notice_message = f"Подписи: {', '.join(map(str, status_list))} не могут быть использованы для подписания документов для отправки ЦХЭД. " \
                             f"Необходимо использовать сертификат выданный ЭМ"
            notice_send(obj, user_id, is_success, notice_message)
            obj.save()
            return False
        is_success = True
        obj.status_id = status_to_id
        notice_message = 'Файл и приложенные подписи успешно отправлены в ЦХЭД'
    else:
        if res_file.text == "Request failed: timeout expire":
            notice_message = 'Сервер ЦХЭД занят. Попробуйте отправить запрос позже.'
        else:
            notice_message = f'Ошибка при отправке файла на сервер: {res_file.text}'
        obj.doc_id = None
    obj.save()
    notice_send(obj, user_id, is_success, notice_message)
    return True


# Функция для обработки значений в JSON-данных
def process_json_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        value = value.replace('nan', '').replace('None', '').replace('null', '').strip()
    return value


# Рекурсивная функция для обработки JSON-данных
def process_json(json_request):
    if isinstance(json_request, dict):
        for key, value in json_request.items():
            json_request[key] = process_json_value(value)
            process_json(json_request[key])
    elif isinstance(json_request, list):
        for i in range(len(json_request)):
            json_request[i] = process_json_value(json_request[i])
            process_json(json_request[i])


def get_obj_responce(model_name, obj_id, status_id, json_request):
    from rest_framework.response import Response
    response = Response()
    model_class, _ = get_model_class(model_name)
    obj = model_class.objects.get(id=obj_id)
    obj.status_id = status_id
    if settings.TEST_MODE:
        response.status_code = 200
        response.content = 'Принято'
    else:
        import requests
        obj.task_id = None
        # Вернуть process_json(json_request), если потребуется обработать json на None
        try:
            response = requests.post(settings.OVGA_URLS['create_pass'], timeout=30, json=json_request)
        except requests.Timeout:
            response.status_code = 500
            response.content = 'Сервер на другой стороне не работает. Попробуйте отправить запрос попозже.'
    return obj, response


@app.task
def send_request_ovga(model_name, obj_id, json_request, status_id, status_to_id, user_id):
    is_success = False
    obj, response = get_obj_responce(model_name, obj_id, status_id, json_request)
    if response.status_code == 200:
        from workflow.models import DocStatus
        status_doc_archived, _ = DocStatus.objects.get_or_create(code='court_order_archived', defaults={'name': 'Отправлен в архив'})
        for pass_req in obj.transport_pass_requests.exclude(status_id=status_doc_archived.id):
            pass_req.status_id = status_to_id
            pass_req.save()
        obj.status_request = 'SUCCESS'
        if hasattr(response, 'content'):
            try:
                obj.body_message_ovga = json.dumps(json.loads(response.content), indent=4, ensure_ascii=False)
            except ValueError:
                obj.body_message_ovga = response.content.decode('utf-8')
        obj.status_id = status_to_id
        is_success = True
        notice_message = obj.body_message_ovga if obj.body_message_ovga else 'Запрос успешно отправлен в ОВГА'
    else:
        obj.status_request = 'ERROR'
        try:
            obj.body_message_ovga = json.dumps(json.loads(response.content), indent=4, ensure_ascii=False)
        except ValueError:
            obj.body_message_ovga = response.content.decode('utf-8')
        notice_message = obj.body_message_ovga if obj.body_message_ovga else 'Ошибка при отправке запроса'
    from workflow.models import LogEvent
    LogEvent.objects.create(name='JSON отправки в ОВГА', event_type='create', data=json_request)
    obj.save()

    notice_send(obj, user_id, is_success, notice_message)
    return True


@app.task
def send_annul_ovga(id_set, user_id, cancel_reason):
    from integrations.ovga_integration.models import TransportPassRequest
    from integrations.ovga_integration.models import PassRequests
    from workflow.models import DocStatus
    is_success = False
    pass_requests = PassRequests.objects.filter(id__in=id_set)
    date_today = datetime.today().strftime('%Y-%m-%dT%H:%M:%S.%fZ')
    status_doc_sucses, _ = DocStatus.objects.get_or_create(code='send_to_annul',
                                                           defaults={'name': 'Отправлено на аннулирование'})
    trans_pass_request = TransportPassRequest(mess_id=uuid.uuid4(), message_date=datetime.now(),
                                              status_request='SUCCESS', status_id=status_doc_sucses.id)
    pass_requests_list = [PassRequests(transport_id=pass_req.transport.id, trans_pass_request=trans_pass_request,
                                  author_id=user_id, pass_series_number=pass_req.pass_series_number,
                                  p_id=uuid.uuid4(), pass_id=pass_req.pass_id, pass_date_cancellation=date_today,
                                  cancel_reason=cancel_reason, status=status_doc_sucses,
                                  request_status_send='SUCCESS') for pass_req in pass_requests]

    pass_request_dicts = [{"requestId": pass_req.p_id.__str__(), "seriesNumber": pass_req.pass_series_number,
                           "passId": str(pass_req.pass_id) if pass_req.pass_id is not None else None,
                           "requestReason": cancel_reason} for pass_req in pass_requests_list]
    json_request = {'annulPassRequests': pass_request_dicts, 'messageId': trans_pass_request.mess_id.__str__(),
                    'messageDate': date_today}
    response_status_code, content = PassRequests.send_annul_request(json_request)
    theme = c.THEME_ERROR
    if response_status_code == 200:
        trans_pass_request.save()
        PassRequests.objects.bulk_create(pass_requests_list)
        theme = c.THEME_SHOW
        is_success = True
        notice_message = 'Запрос на аннулирование заявок отправлен в ОВГА'
    elif content == "Request failed: timeout expire":
        notice_message = 'Сервер на другой стороне не может принять запрос. Попробуйте выполнить запрос позже.'
    else:
        notice_message = f'Ошибка сервера: {content}'
    notice_send(trans_pass_request, user_id, is_success, notice_message, theme)
    return True


@app.task
def change_stat_model_task(model_name, id_set, t_id, doc_type_code, req_user_id):
    from model_app.billing.billing_utils import send_error_notice
    from workflow.tools.base import Transition
    from workflow.tools.base import TransitionException
    trans_docs_model = ContentType.objects.filter(model=model_name).first().model_class()
    transport_documents = trans_docs_model.objects.filter(id__in=id_set)
    doc_type = DocType.objects.get(code=doc_type_code)
    req_user = User.objects.get(id=req_user_id)
    content = ''
    content_error = ''
    for transport_document in transport_documents:
        try:
            Transition.run_transition(transport_document, req_user, trans_docs_model, doc_type, t_id)
            content += f'{transport_document.id}, '
        except TransitionException as e:
            content_error += f'Переход с ID: {transport_document.id} не выполнен. Ошибка: {e.args.__str__()} \r\n'
    if content:
        notice_send(None, req_user_id, True, f"Выполнен переход с ID: {content}", theme=c.THEME_SUCCESS)
    if content_error:
        send_error_notice(req_user, content_error)
    return True


@app.task
def generate_xslx_ovga(model_name: str, id_set: list, req_user_id: int):
    trans_pass_request_model = ContentType.objects.filter(model=model_name).first().model_class()
    for id_trans_pass_request in id_set:
        pass_requests = PassRequests.objects.filter(trans_pass_request=id_trans_pass_request)
        transport_ids = pass_requests.values_list('transport_id', flat=True)
        transport = Transport.objects.filter(id__in=transport_ids)

        # Получаем ID типа документа "Водительское удостоверение"
        type_document_ids = TypeDocument.objects.filter(type_document='Водительское удостоверение').values_list(
            'id', flat=True)

        # Создаем подзапрос для агрегации серий и номеров документов водителей в строку
        drivers_document_subquery = Subquery(
            DriverDocument.objects.filter(
                driver__transports__id=OuterRef('id'),
                types_of_document_id__in=type_document_ids
            ).annotate(
                doc_info=Concat('series', Value(' '), 'number', output_field=CharField())
            ).values('doc_info').order_by('doc_info').annotate(
                docs=StringAgg('doc_info', delimiter=', ')
            ).values('docs')[:1],
            output_field=CharField()
        )

        # Добавляем аннотации для телефона и адреса участника
        phone = ContactType.objects.filter(code='phone').first()
        address = ContactType.objects.filter(code='address').first()

        participant_phone = Subquery(
            Contact.objects.filter(participant_id=OuterRef('participant_id'), contact_type=phone)
            .annotate(phone=StringAgg('value', delimiter=', '))
            .values('phone')[:1], output_field=CharField())

        participant_address = Subquery(
            Contact.objects.filter(participant_id=OuterRef('participant_id'), contact_type=address)
            .annotate(addresses=StringAgg('value', delimiter=', '))
            .values('addresses')[:1], output_field=CharField())

        # Аннотируем QuerySet транспорта информацией о документах водителей, телефоне и адресе участника
        transport = transport.annotate(
            brand_model=Subquery(
                TransportDocument.objects.filter(transport__id=OuterRef('id'))
                .exclude(transport_brand_pts__isnull=True)
                .values('transport_brand_pts')[:1], output_field=CharField()),
            type_name=Subquery(
                TransportDocument.objects.filter(transport__id=OuterRef('id'))
                .exclude(type_of_transport__isnull=True)
                .values('type_of_transport')[:1], output_field=CharField()),
            driver_documents=drivers_document_subquery,
            participant_phone=participant_phone,
            participant_address=participant_address,
            ogrn_inn=Concat('participant__ogrn', Value(' / '), 'participant__inn', output_field=CharField())
        ).values(
            'number', 'type_name', 'brand_model', 'participant__full_name', 'participant_address',
            'participant_phone', 'driver_documents', 'ogrn_inn'
        )
        df = pd.DataFrame(transport)
        df.fillna('', inplace=True)
        df = df.rename(columns={
            'number': 'Государственный регистрационный знак',
            'type_name': 'Тип транспортного средства',
            'brand_model': 'Марка, модель транспортного средства',
            'participant__full_name': 'Полное наименование юридического лица/индивидуального предпринимателя',
            'participant_address': 'Почтовый адрес юридического лица/индивидуального предпринимателя',
            'participant_phone': 'Телефон юридического лица/индивидуального предпринимателя',
            'driver_documents': 'Серия, номер водительского удостоверения водителя транспортного средства',
            'ogrn_inn': 'ОГРН/ОГРНИП/ИНН юридического лица/индивидуального предпринимателя/физического лица'
        })
        filename = f'Массовый запрос {id_trans_pass_request}.xlsx'
        sheet_name = 'Данные по пропускам'
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
        output.seek(0)
        req_user = User.objects.get(id=req_user_id)
        up_file = UploadedFile.create(filename, output.read(), req_user)
        if trans_pass_request_model:
            trans_pass_request = trans_pass_request_model.objects.get(id=id_trans_pass_request)
            trans_pass_request.file = up_file
            trans_pass_request.save()
    notice_send(None, req_user_id, True, f"Сформированы файлы по выбранным массовым запросам", theme=c.THEME_SUCCESS)
    return True
